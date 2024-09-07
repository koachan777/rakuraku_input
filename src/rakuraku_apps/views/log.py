import io
import base64
from datetime import datetime, timedelta
import numpy as np

from django.views.generic import TemplateView
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse

from rakuraku_apps.models import ShrimpModel, WaterQualityModel, TankModel, ShrimpModel
from rakuraku_apps.forms.log import WaterQualityEditForm

import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter, DayLocator
import japanize_matplotlib
import openpyxl
from openpyxl.utils import get_column_letter


class TableOrGraphView(TemplateView):
    template_name = 'log/table_or_graph.html'



class TableView(TemplateView):
    template_name = 'log/table.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        shrimp_id = self.request.GET.get('shrimp')
        item = self.request.GET.get('item')

        if not start_date:
            start_date = (datetime.today() - timedelta(days=7)).strftime('%Y-%m-%d')
        if not end_date:
            end_date = datetime.today().strftime('%Y-%m-%d')

        query = Q(date__range=[start_date, end_date])

        if shrimp_id:
            query &= Q(tank__shrimp__id=shrimp_id)

        water_quality_data = WaterQualityModel.objects.filter(query).select_related('tank', 'tank__shrimp')

        if item:
            water_quality_data = water_quality_data.values('date', 'tank__name', 'tank__id', item).order_by('date', 'tank__id')
            water_quality_data_by_date = {}
            for data in water_quality_data:
                date = data['date']
                tank_name = data['tank__name']
                value = data[item]
                if date not in water_quality_data_by_date:
                    water_quality_data_by_date[date] = {}
                water_quality_data_by_date[date][tank_name] = value
            context['water_quality_data_by_date'] = dict(sorted(water_quality_data_by_date.items()))
    
            if shrimp_id:
                context['tanks'] = TankModel.objects.filter(shrimp__id=shrimp_id, water_quality__date__range=[start_date, end_date]).distinct().order_by('id')
            else:
                context['tanks'] = TankModel.objects.filter(water_quality__date__range=[start_date, end_date]).distinct().order_by('id')

        else:
            water_quality_data = water_quality_data.values('date', 'tank__name', 'pH', 'DO', 'salinity', 'NH4', 'NO2', 'NO3', 'Ca', 'Al', 'Mg', 'water_temperature', 'room_temperature', 'notes')

        # 水槽名の昇順と日付の昇順で並び替え
        water_quality_data = water_quality_data.order_by('tank__name', 'date').values('id', 'date', 'tank__name', 'pH', 'DO', 'salinity', 'NH4', 'NO2', 'NO3', 'Ca', 'Al', 'Mg', 'water_temperature', 'room_temperature', 'notes')
        context['water_quality_data'] = water_quality_data
        context['shrimps'] = ShrimpModel.objects.all()
        context['start_date'] = start_date
        context['end_date'] = end_date

        return context
    
def edit_water_quality(request, pk):
    water_quality = get_object_or_404(WaterQualityModel, pk=pk)
    if request.method == 'POST':
        form = WaterQualityEditForm(request.POST, instance=water_quality)
        if form.is_valid():
            form.save()
            return redirect('rakuraku_apps:table')
    else:
        form = WaterQualityEditForm(instance=water_quality)
    return render(request, 'log/edit_water_quality.html', {'form': form})

def delete_water_quality(request, pk):
    water_quality = get_object_or_404(WaterQualityModel, pk=pk)
    if request.method == 'POST':
        water_quality.delete()
    return redirect('rakuraku_apps:table')

def export_to_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    shrimp_id = request.GET.get('shrimp')
    item = request.GET.get('item')

    if not start_date:
        start_date = (datetime.today() - timedelta(days=7)).strftime('%Y-%m-%d')
    if not end_date:
        end_date = datetime.today().strftime('%Y-%m-%d')

    query = Q(date__range=[start_date, end_date])

    if shrimp_id:
        tanks = TankModel.objects.filter(shrimp__id=shrimp_id, water_quality__date__range=[start_date, end_date]).distinct().order_by('id')
    else:
        tanks = TankModel.objects.filter(water_quality__date__range=[start_date, end_date]).distinct().order_by('id')

    water_quality_data = WaterQualityModel.objects.filter(query).select_related('tank', 'tank__shrimp')

    # Excelファイルを作成
    wb = openpyxl.Workbook()
    ws = wb.active

    if item:
        # 項目名を日本語に変換
        item_name_dict = {
            'water_temperature': '水温',
            'pH': 'pH',
            'DO': 'DO',
            'salinity': '塩分濃度',
            'NH4': 'NH4',
            'NO2': 'NO2',
            'NO3': 'NO3',
            'Ca': 'Ca',
            'Al': 'Al',
            'Mg': 'Mg',
        }
        item_name = item_name_dict.get(item, item)

        # 項目名をタイトルとして一番左に表示
        ws.cell(row=1, column=1, value=item_name)

        # ヘッダー行を書き込む
        col = 2
        for tank in tanks:
            ws.cell(row=2, column=col, value=tank.name)
            col += 1

        # データ行を書き込む
        dates = water_quality_data.values_list('date', flat=True).distinct().order_by('date')
        row = 3
        for date in dates:
            ws.cell(row=row, column=1, value=date)
            col = 2
            for tank in tanks:
                tank_data = water_quality_data.filter(date=date, tank=tank).first()
                if tank_data:
                    ws.cell(row=row, column=col, value=getattr(tank_data, item) or '-')
                col += 1
            row += 1
    else:
        # ヘッダー行を書き込む
        ws.cell(row=1, column=1, value='日付')
        col = 2
        for tank in tanks:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+12)
            ws.cell(row=1, column=col, value=tank.name)
            col += 13

        items = ['室温', '水温', 'pH', 'DO', '塩分濃度', 'NH4', 'NO2', 'NO3', 'Ca', 'Al', 'Mg', '備考']
        row = 2
        col = 1
        ws.cell(row=row, column=col, value='')
        col += 1
        for tank in tanks:
            for item_name in items:
                ws.cell(row=row, column=col, value=item_name)
                col += 1
            col += 1

        # データ行を書き込む
        dates = water_quality_data.values_list('date', flat=True).distinct().order_by('date')
        row = 3
        for date in dates:
            ws.cell(row=row, column=1, value=date)
            col = 2
            for tank in tanks:
                tank_data = water_quality_data.filter(date=date, tank=tank).first()
                if tank_data:
                    ws.cell(row=row, column=col, value=tank_data.room_temperature or '-')
                    ws.cell(row=row, column=col+1, value=tank_data.water_temperature or '-')
                    ws.cell(row=row, column=col+2, value=tank_data.pH or '-')
                    ws.cell(row=row, column=col+3, value=tank_data.DO or '-')
                    ws.cell(row=row, column=col+4, value=tank_data.salinity or '-')
                    ws.cell(row=row, column=col+5, value=tank_data.NH4 or '-')
                    ws.cell(row=row, column=col+6, value=tank_data.NO2 or '-')
                    ws.cell(row=row, column=col+7, value=tank_data.NO3 or '-')
                    ws.cell(row=row, column=col+8, value=tank_data.Ca or '-')
                    ws.cell(row=row, column=col+9, value=tank_data.Al or '-')
                    ws.cell(row=row, column=col+10, value=tank_data.Mg or '-')
                    ws.cell(row=row, column=col+11, value=tank_data.notes or '-')
                col += 13
            row += 1

    # セルの幅を自動調整
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, sum(2 if ord(c) > 127 else 1 for c in str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

    # Excelファイルをレスポンスとして返す
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=water_quality_data.xlsx'
    wb.save(response)
    return response



class GraphView(TemplateView):
    template_name = 'log/graph.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        shrimp_id = self.request.GET.get('shrimp')
        item = self.request.GET.get('item')
        compare_last_year = self.request.GET.get('compare_last_year')

        if not start_date:
            start_date = (datetime.today() - timedelta(days=7)).strftime('%Y-%m-%d')
        if not end_date:
            end_date = datetime.today().strftime('%Y-%m-%d')

        item_labels = {
            'water_temperature': '水温',
            'pH': 'pH',
            'DO': 'DO',
            'salinity': '塩分濃度',
            'NH4': 'NH4',
            'NO2': 'NO2',
            'NO3': 'NO3',
            'Ca': 'Ca',
            'Al': 'Al',
            'Mg': 'Mg'
        }

        query = Q()
        if start_date and end_date:
            query &= Q(date__range=[start_date, end_date])
        if shrimp_id:
            query &= Q(tank__shrimp__id=shrimp_id)

        water_quality_data = WaterQualityModel.objects.filter(query).select_related('tank', 'tank__shrimp')

        if item:
            water_quality_data = water_quality_data.values('date', 'tank__name', 'tank__shrimp__id', item)
        else:
            item = 'water_temperature'  # デフォルトの項目をwater_temperatureに設定
            water_quality_data = water_quality_data.values('date', 'tank__name', 'tank__shrimp__id', item)

        water_quality_data = water_quality_data.order_by('date', 'tank__id')

        # 現在の期間のグラフを描画
        current_graph = draw_graph(water_quality_data, item, item_labels, start_date, end_date, compare_last_year)
        
        last_year_graph = None
        last_year_data_exists = False
        if compare_last_year:
            # 一年前の期間を計算
            last_year_start_date = (datetime.strptime(start_date, '%Y-%m-%d') - timedelta(days=365)).strftime('%Y-%m-%d')
            last_year_end_date = (datetime.strptime(end_date, '%Y-%m-%d') - timedelta(days=365)).strftime('%Y-%m-%d')

            # 一年前のデータを取得
            last_year_query = Q(date__range=[last_year_start_date, last_year_end_date])
            if shrimp_id:
                last_year_query &= Q(tank__shrimp__id=shrimp_id)
            last_year_water_quality_data = WaterQualityModel.objects.filter(last_year_query).select_related('tank', 'tank__shrimp')
            last_year_water_quality_data = last_year_water_quality_data.values('date', 'tank__name', 'tank__shrimp__id', item)
            last_year_water_quality_data = last_year_water_quality_data.order_by('date', 'tank__id')

            if last_year_water_quality_data.exists() and any(data[item] is not None for data in last_year_water_quality_data):
                last_year_data_exists = True
                # 一年前のグラフを描画
                last_year_graph = draw_graph(last_year_water_quality_data, item, item_labels, last_year_start_date, last_year_end_date, compare_last_year)

        context['current_graph'] = current_graph
        context['last_year_graph'] = last_year_graph
        context['last_year_data_exists'] = last_year_data_exists
        context['shrimps'] = ShrimpModel.objects.all()
        context['selected_item'] = item
        context['start_date'] = start_date
        context['end_date'] = end_date
        context['compare_last_year'] = compare_last_year
        return context

def draw_graph(water_quality_data, item, item_labels, start_date, end_date, compare_last_year=False):
    fig, ax = plt.subplots(figsize=(10, 6), tight_layout=True)
    fig.subplots_adjust(top=0.9)
    fig.suptitle(item_labels.get(item, ''), fontsize=24, fontweight='bold')

    # 系統ごとに水槽をグルーピング
    shrimp_tanks = {}
    for tank_data in water_quality_data:
        shrimp_id = tank_data.get('tank__shrimp__id')
        if shrimp_id:
            if shrimp_id not in shrimp_tanks:
                shrimp_tanks[shrimp_id] = []
            shrimp_tanks[shrimp_id].append(tank_data)

    colors = plt.cm.get_cmap('tab20', sum(len(set(tank['tank__name'] for tank in tanks)) for tanks in shrimp_tanks.values()))
    color_index = 0

    for shrimp_id, tanks in shrimp_tanks.items():
        # 系統内の水槽をID順にソート
        tanks.sort(key=lambda x: x['tank__name'])

        # 重複する水槽名を解消
        unique_tank_names = sorted(set(tank['tank__name'] for tank in tanks))

        for tank_name in unique_tank_names:
            dates = []
            values = []
            for data in water_quality_data:
                if data['tank__name'] == tank_name:
                    dates.append(data['date'])
                    if data[item] is None:
                        values.append(None)
                    else:
                        values.append(data[item])
            color = colors(color_index)
            mask = [value is not None for value in values]
            ax.plot([date for date, m in zip(dates, mask) if m], [value for value, m in zip(values, mask) if m], marker='o', label=tank_name, color=color, linewidth=2)
            color_index += 1


    handles, labels = ax.get_legend_handles_labels()
    ax.legend(handles, labels, loc='upper right')

    # 日付の間隔を計算
    num_days = (datetime.strptime(end_date, '%Y-%m-%d') - datetime.strptime(start_date, '%Y-%m-%d')).days + 1
    if num_days > 20:
        interval = 5
    else:
        interval = 1

    # 「昨年のデータと比較する」チェックボックスがオンの場合は、常に年を含める
    if compare_last_year:
        date_format = '%Y/%m/%d'
    else:
        # 年をまたぐかどうかを判定
        start_year = datetime.strptime(start_date, '%Y-%m-%d').year
        end_year = datetime.strptime(end_date, '%Y-%m-%d').year
        if start_year != end_year:
            date_format = '%Y/%m/%d'
        else:
            date_format = '%m/%d'

    ax.xaxis.set_major_formatter(DateFormatter(date_format))
    ax.xaxis.set_major_locator(DayLocator(interval=interval))
    fig.autofmt_xdate()
    ax.grid(axis='y', linestyle='-', linewidth=0.5, color='gray', alpha=0.7)

    # 単位を設定
    unit = ''
    if item == 'water_temperature':
        unit = '°C'
    elif item == 'DO':
        unit = 'mg/L'
    elif item == 'salinity':
        unit = '%'
    elif item in ['NH4', 'NO2', 'NO3', 'Ca', 'Al', 'Mg']:
        unit = 'mg/L'

    # 単位がある場合のみ、グラフの枠の外の左上に表示
    if unit:
        plt.figtext(0.03, 0.925, f'({unit})', fontsize=12, bbox=dict(facecolor='white', alpha=0.8, edgecolor='none', pad=2))

    # グラフをbase64エンコードされた文字列に変換
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    image_png = buffer.getvalue()
    graph = base64.b64encode(image_png).decode('utf-8')
    buffer.close()

    return graph
